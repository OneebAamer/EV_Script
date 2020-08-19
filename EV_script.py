#EV algorithm
#------------------------------#
#Importing Modules
import csv
from openpyxl import load_workbook
from math import ceil, floor, log
#Defining object Vehicle, which will be used for each vehicle we test
class Vehicle:
    def __init__(self, name, price, distpcharge, chargeTime, max_load,battery,weight):
        self.name = name
        self.price = price
        self.distpcharge = distpcharge
        self.actual_distpcharge = distpcharge * 0.85
        self.chargeTime = chargeTime
        self.max_load = max_load
        self.battery = battery
        self.efficiency = self.actual_distpcharge/self.battery
        self.weight = weight
#Define Global Variables
sort_list = []
vehicle_list = []
scores = []
num_of_cars = 0
#Define subroutine for comparing vehicles in a list.
def compare(vehicleArray):
    for i in range(0,len(vehicleArray)):
            for j in range(i,len(vehicleArray)):
                if i == j:
                    pass
                else:
                    distpcharge = vehicleArray[i].actual_distpcharge/(vehicleArray[j].actual_distpcharge + vehicleArray[i].actual_distpcharge)
                    load = vehicleArray[i].max_load/(vehicleArray[j].max_load + vehicleArray[i].max_load)
                    price = vehicleArray[i].price/(vehicleArray[j].price + vehicleArray[i].price)
                    chargeTime = vehicleArray[i].chargeTime/(vehicleArray[j].chargeTime + vehicleArray[i].chargeTime)
                    efficiency = vehicleArray[i].efficiency/(vehicleArray[i].efficiency + vehicleArray[j].efficiency)
                    score = distpcharge + load + efficiency - price - chargeTime
                    if score > 0.5:
                        scores[i] += 1
                    elif score < 0.5:
                        scores[j] += 1
                    else:
                        pass
#Announce the Winner using the data given.
def winner(score_list,n):
    wb = load_workbook(filename = 'results.xlsx')
    ws = wb.active
    ws.cell(row = 14,column = 1).value = "Position:"
    ws.cell(row = 14,column = 2).value = "Name:"
    print("Positions: ")
    for i in range(0,len(score_list)):
        sort_list[i][0] = vehicle_list[i].name
        sort_list[i][1] = n - score_list[i]
    sort_list.sort(key = lambda x: x[1])
    for i in range(0,len(score_list)):
        name = sort_list[i][0]
        position = sort_list[i][1]
        print(name,":",position)
        ws.cell(row=i+15,column=1).value = sort_list[i][1]
        ws.cell(row=i+15,column=2).value = sort_list[i][0]
    try:
        wb.save('results.xlsx')
    except:
        print("Unable to open file to save data. Is the file open?")
#Draw results for visuals into a new document
def results(vehicle_list):
    index = 1
    wb = load_workbook(filename = 'results.xlsx')
    ws = wb.active
    ws.cell(row = 1,column = 1).value = 'Name:'
    ws.cell(row = 1,column = 2).value = 'Time (Hours):'
    ws.cell(row = 1,column = 3).value = 'CO2 (KG):'
    const_travelTime = 526/58
    for vehicle in vehicle_list:
        chargingHours = ceil(526/vehicle.actual_distpcharge) * vehicle.chargeTime
        totalTime = chargingHours + const_travelTime
        emission = 0
        if vehicle.weight < 1.3:
            emission = 0.06766
        elif vehicle.weight > 1.3 and vehicle.weight < 1.74:
            emission = 0.09234
        elif vehicle.weight > 1.74 and vehicle.weight < 3.5:
            emission = 0.13453
        carbon = emission * 526
        ws.cell(row=index+1,column=1).value = vehicle.name
        ws.cell(row=index+1,column=2).value = totalTime
        ws.cell(row=index+1,column=3).value = carbon
        index += 1
    try:
        wb.save('results.xlsx')
        print("Data has been updated in results.csv")
    except:
        print("Unable to open file to save data. Is the file open?")
#Subroutine called when inputting data is chosen
def chooseInput():
    while num_of_cars < 2 or num_of_cars > 10:
        num_of_cars = int(input("How many cars will you be comparing?"))
    for i in range(0,num_of_cars):
        scores.append(0)
    for i in range(0,num_of_cars):
        sort_list.append([0,0])
        print("Enter the name of car",i + 1,": ")
        name = input()
        price = int(input("Enter the price: "))
        distpcharge = int(input("Enter the distance it can travel per charge: "))
        size = int(input("Enter the charging time for the vehicle: "))
        max_load = int(input("Enter the maximum load it can carry: "))
        battery = float(input("Enter the battery capacity of the vehicle: "))
        weight = float(input("Enter the weight of the vehicle: "))
        add_vehicle = Vehicle(name,price,distpcharge,chargeTime,max_load,battery,weight)
        vehicle_list.append(add_vehicle)
        print("\n")
        compare(vehicle_list)
        winner(scores,num_of_cars)
        results(vehicle_list)
#Subroutine called when CSV file is used as input
def chooseCsv():
    cars = 0
    with open('data.csv','r', newline='') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            if row[0] == 'END':
                break
            elif row[0] != 'Car Name':
                sort_list.append([0,0])
                scores.append(0)
                cars += 1
                name = row[0]
                price = float(row[1])
                distpcharge = float(row[2])
                chargeTime = float(row[3])
                load = float(row[4])
                battery = float(row[5])
                weight = float(row[6])
                add_vehicle = Vehicle(name,price,distpcharge,chargeTime,load,battery,weight)
                vehicle_list.append(add_vehicle)
    compare(vehicle_list)
    winner(scores,cars)
    results(vehicle_list)
#Start of the program.
def start():
    ans = input("Would you like to use csv or input?: ").lower()
    if ans == "csv":
        chooseCsv()
    elif ans == "input":
        chooseInput()
    else:
        start()
start()






            
    

